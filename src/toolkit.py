import numpy as np
import torch
import re
import openpyxl, xlrd
import numpy as np

def get_table_vec(vectarr, reduce=False):
    n,m,d = vectarr.shape
    row_med = np.stack([np.median(vectarr, axis=1)]*m, axis=1)
    row_mean = np.stack([np.mean(vectarr, axis=1)]*m, axis=1)
    col_med = np.stack([np.median(vectarr, axis=0)]*n, axis=0)
    col_mean = np.stack([np.mean(vectarr, axis=0)]*n, axis=0)
    
    table_med = np.median(vectarr, axis=[0,1])
    table_med = np.stack([table_med]*m, 0)
    table_med = np.stack([table_med]*n, 0)
    
    table_mean = np.mean(vectarr, axis=0)
    table_mean = np.mean(table_mean, axis=0)
    table_mean = np.stack([table_mean]*m, 0)
    table_mean = np.stack([table_mean]*n, 0)
    
    row_dev_med = np.mean((vectarr - row_med)**2, axis=0)
    row_dev_med = np.mean(row_dev_med, axis=0)
    col_dev_med = np.mean((vectarr - col_med)**2, axis=0)
    col_dev_med = np.mean(col_dev_med, axis=0)
    table_dev_med = np.mean((vectarr - table_med)**2, axis=0)
    table_dev_med = np.mean(table_dev_med, axis=0)
    
    row_dev_mean = np.mean((vectarr - row_mean)**2, axis=0)
    row_dev_mean = np.mean(row_dev_mean, axis=0)
    col_dev_mean = np.mean((vectarr - col_mean)**2, axis=0)
    col_dev_mean = np.mean(col_dev_mean, axis=0)
    table_dev_mean = np.mean((vectarr - table_mean)**2, axis=0)
    table_dev_mean = np.mean(table_dev_mean, axis=0)
    if not reduce:
        return np.hstack([row_dev_med,col_dev_med,table_dev_med,row_dev_mean,col_dev_mean,table_dev_mean])
    else:
        reduce = lambda x: np.sqrt(np.sum(x**2))
        return np.hstack([reduce(row_dev_med),reduce(col_dev_med),reduce(table_dev_med),reduce(row_dev_mean),reduce(col_dev_mean),reduce(table_dev_mean)])

def get_vec_tarr(t, ce):
    def get_minimized_cell(tarr, rind, cind, n, m):
        text = tarr[rind][cind].strip()
        if text == '' or text is None:
            text = '__NULL__'
        r0 = tarr[rind][cind+1].strip() if cind < m-1 else '__RIGHT__'
        l0 = tarr[rind][cind-1].strip() if cind > 0 else '__LEFT__'
        t0 = tarr[rind-1][cind].strip() if rind > 0 else '__TOP__'
        b0 = tarr[rind+1][cind].strip() if rind < n-1 else '__BOTTOM__'

        r1 = tarr[rind][cind+2].strip() if cind < m-2 else '__RIGHT__'
        l1 = tarr[rind][cind-2].strip() if cind > 1 else '__LEFT__'
        t1 = tarr[rind-2][cind].strip() if rind > 1 else '__TOP__'
        b1 = tarr[rind+2][cind].strip() if rind < n-2 else '__BOTTOM__'

        if r0 == '' or r0 is None: r0 = '__NULL__'
        if l0 == '' or l0 is None: l0 = '__NULL__'
        if t0 == '' or t0 is None: t0 = '__NULL__'
        if b0 == '' or b0 is None: b0 = '__NULL__'
        if r1 == '' or r1 is None: r1 = '__NULL__'
        if l1 == '' or l1 is None: l1 = '__NULL__'
        if t1 == '' or t1 is None: t1 = '__NULL__'
        if b1 == '' or b1 is None: b1 = '__NULL__'
        res = dict(text=text,
                 rightText=r0,
                 leftText=l0,
                 topText=t0,
                 bottomText=b0,
                 rightText1=r1,
                 leftText1=l1,
                 topText1=t1,
                 bottomText1=b1)
        return res

    class myiter:
        def __init__(self, arr):
            self.it = iter(arr)

        def next(self):
            return next(self.it)

    n,m = len(t) , len(t[0])

    vec_tarr = np.zeros([n,m, ce.encdim], dtype=float)

    for i in range(n):
        row = []
        for j in range(m):
            c = get_minimized_cell(t, i, j, n, m)
            row.append(c)
        with torch.no_grad():
            row_vecs = ce.transform(myiter(row), m, 1, m)
        vec_tarr[i,:,:] = row_vecs
    return vec_tarr

def calc_table_vector(tarr, ce, reduce=False, mode='cbow'):
    vtarr = get_vec_tarr(tarr, ce, mode)
    tvec = get_table_vec(vtarr, reduce)
    return tvec

def create_fingerprint(table):
    table = str(table)
    all_tokens = list(set(re.split('[^\w]+',table)))
    all_tokens = sorted(all_tokens)
    fingerprint = '-'.join(all_tokens)
    return fingerprint


def get_excel_index_as_ints(r):
    col_names = [x for x in string.ascii_lowercase]
    col_names += ['{}{}'.format(x,y) for x in string.ascii_lowercase for y in string.ascii_lowercase]
    col_names += ['{}{}{}'.format(x,y,z) for x in string.ascii_lowercase for y in string.ascii_lowercase for z in string.ascii_lowercase]
    
    if ':' in r:
        rl, rr = r.split(':')
        rl_r, rl_c = int(re.sub('[a-z]', '', rl))-1 , col_names.index(re.sub('[0-9]', '', rl))
        rr_r, rr_c = int(re.sub('[a-z]', '', rr))-1 , col_names.index(re.sub('[0-9]', '', rr))
        i,j = list(range(rl_r, rr_r+1)), list(range(rl_c, rr_c+1))
        ij = list(product(i,j))
        i = [x[0] for x in ij]
        j = [x[1] for x in ij]
    else:    
        i = [int(re.sub('[a-z]', '', r))-1]
        j = [col_names.index(re.sub('[0-9]', '', r))]
        
    return i,j

def get_excel_range_as_slice(tot_range, block_range):
    col_names = [x for x in string.ascii_lowercase]
    col_names += ['{}{}'.format(x,y) for x in string.ascii_lowercase for y in string.ascii_lowercase]
    col_names += ['{}{}{}'.format(x,y,z) for x in string.ascii_lowercase for y in string.ascii_lowercase for z in string.ascii_lowercase]
    
    
    
    if ':' in tot_range:
        trl, trr = tot_range.split(':')
    else:
        trl = trr = tot_range
        
    tr_tr, tr_br, tr_lc, tr_rc = int(re.sub('[a-z]', '', trl)), int(re.sub('[a-z]', '', trr)), re.sub('[0-9]', '', trl), re.sub('[0-9]', '', trr)
    
    if block_range is None:
        return '{}:{},{}:{}'.format(tr_tr-1, tr_br, col_names.index(tr_lc),col_names.index(tr_rc)+1) 
        
    if ':' in block_range:
        brl, brr = block_range.split(':')
    else:
        brl = brr = block_range
    
    br_tr, br_br, br_lc, br_rc = int(re.sub('[a-z]', '', brl)), int(re.sub('[a-z]', '', brr)), re.sub('[0-9]', '', brl), re.sub('[0-9]', '', brr)
    
    col_slice = '{}:{}'.format(col_names.index(br_lc)-col_names.index(tr_lc), col_names.index(br_rc)-col_names.index(tr_lc)+1)
    row_slice = '{}:{}'.format(br_tr-tr_tr, br_br-tr_tr+1)
    return '{},{}'.format(row_slice, col_slice) 

def get_excel_range_dimension(r):
    col_names = [x for x in string.ascii_lowercase]
    col_names += ['{}{}'.format(x,y) for x in string.ascii_lowercase for y in string.ascii_lowercase]
    col_names += ['{}{}{}'.format(x,y,z) for x in string.ascii_lowercase for y in string.ascii_lowercase for z in string.ascii_lowercase]
    
    if ':' in r:
        rl, rr = r.split(':')
    else:
        return (1,1)
    lrow = int(re.sub('[a-z]', '', rl))
    rrow = int(re.sub('[a-z]', '', rr))
    lcol = col_names.index(re.sub('[0-9]', '', rl))
    rcol = col_names.index(re.sub('[0-9]', '', rr))
    
    return (rrow-lrow+1, rcol-lcol+1)

def get_sheet_names(fpath, file_type='xlsx'):
    res = []
    if file_type == 'xlsx':
        book = openpyxl.load_workbook(fpath, read_only=False, data_only=True)
        sheet_names = book.sheetnames
    else:
        book = xlrd.open_workbook(fpath)
        sheet_names = book.sheet_names()
    return sheet_names

def get_sheet_tarr(fpath, sname, file_type='xlsx'):
    res = []
    
    if file_type == 'xlsx':
        book = openpyxl.load_workbook(fpath, read_only=False, data_only=True)
        sheet = book[sname]
        sheet_array = np.array(list(sheet.values))
    else:
        book = xlrd.open_workbook(fpath)
        sheet = book.sheet_by_name(sname)
        N, M = sheet.nrows, sheet.ncols
        sheet_array = np.empty([N,M], dtype=object)
        for ri in range(N):
            rvals = sheet.row_values(ri)
            sheet_array[ri] = rvals
    vf = np.vectorize(str) 
    sheet_array = vf(sheet_array)
    return sheet_array